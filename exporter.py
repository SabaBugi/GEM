import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER
from database import DB_NAME, get_images


# ---------------- FONT SETUP ----------------
# Make sure you have Noto Sans Georgian fonts in 'fonts' folder
pdfmetrics.registerFont(TTFont("NotoSansGeorgian", "fonts/NotoSansGeorgian-Regular.ttf"))
pdfmetrics.registerFont(TTFont("NotoSansGeorgian-Bold", "fonts/NotoSansGeorgian-Bold.ttf"))

label_style = ParagraphStyle(name="Label", fontName="NotoSansGeorgian-Bold", fontSize=11, leading=14, alignment=TA_CENTER)
value_style = ParagraphStyle(name="Value", fontName="NotoSansGeorgian", fontSize=11, leading=14)


# ---------------- HEADERS ----------------
HEADERS = [
    "ID", "კოდი", "ნივთი", "კატეგორია", "აღმოჩენის ადგილი", "აღწერა",
    "პერიოდი", "მდებარეობა", "მდგომარეობა", "სტატუსი", "კურატორი", "თარიღი"
]

def get_all_artefacts():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, artefact_code, name, category, origin, description,
               period, location, condition, status, curator, date_added
        FROM artefacts
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

# ---------------- EXCEL EXPORT ----------------
def export_to_excel(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artefacts"

    # Headers
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    rows = get_all_artefacts()
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            text = str(value) if value else ""
            wrap = True if " " in text else False
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    # ---------------- MANUAL COLUMN WIDTHS ----------------
    manual_widths = {
        "ID": 5,
        "კოდი": 12,
        "ნივთი": 20,
        "კატეგორია": 18,
        "აღმოჩენის ადგილი": 25,
        "აღწერა": 30,
        "პერიოდი": 15,
        "მდებარეობა": 20,
        "მდგომარეობა": 18,
        "სტატუსი": 18,
        "კურატორი": 20,
        "თარიღი": 11
    }

    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = manual_widths.get(header, 15)

    # ---------------- ROW HEIGHTS ----------------
    for row_idx in range(2, ws.max_row + 1):
        max_height = 15
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                text = str(cell.value)
                if " " in text:
                    est_lines = (len(text) // 20) + 1  # rough estimate for line breaks
                    height = est_lines * 15
                    max_height = max(max_height, height)
        ws.row_dimensions[row_idx].height = max_height

    wb.save(filename)
    print(f"✅ Exported to Excel: {filename}")


# ---------------- PDF EXPORT ----------------
def export_to_pdf(filename):
    doc = SimpleDocTemplate(filename, pagesize=A4,
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    artefacts = get_all_artefacts()
    page_width, page_height = A4
    table_width = page_width - 60

    def build_table(artefact):
        artefact_id = artefact[0]
        description = artefact[5] or ""
        images = get_images(artefact_id)

        fields = [
            ("კოდი", artefact[1]),
            ("ნივთი", artefact[2]),
            ("კატეგორია", artefact[3]),
            ("აღმოჩენის ადგილი", artefact[4]),
            ("პერიოდი", artefact[6]),
            ("მდებარეობა", artefact[7]),
            ("მდგომარეობა", artefact[8]),
            ("სტატუსი", artefact[9]),
            ("აღწერა", description),
        ]

        # --- Prepare table data ---
        table_data = []
        for label, value in fields:
            # Labels are bold
            label_para = Paragraph(label, label_style)
            # Values are regular (non-bold)
            value_para = Paragraph(str(value) if value else "", value_style)
            
            # For description, use value_style
            if label == "აღწერა":
                row = [label_para, value_para, ""]
            else:
                row = [label_para, value_para, ""]
            
            table_data.append(row)

        # --- Row heights ---
        row_heights = []
        for label, _ in fields:
            row_heights.append(32 if label != "აღწერა" else 64)


        # --- PHOTO header ---
        table_data[0][2] = Paragraph("<b>ფოტო</b>", label_style)

        # --- Image ---
        img_obj = None
        if images and os.path.exists(images[0]):
            try:
                img_obj = Image(images[0])
                orig_width, orig_height = img_obj.imageWidth, img_obj.imageHeight
                
                # reserve 10% padding on all sides
                reserved_width = table_width * 0.3 * 0.9  # 90% of column width
                reserved_height = 180 * 0.9              # 90% of row height
                
                scale = min(reserved_width / orig_width, reserved_height / orig_height)
                img_obj.drawWidth = orig_width * scale
                img_obj.drawHeight = orig_height * scale
                img_obj.hAlign = "CENTER"
            except:
                img_obj = None

        if img_obj:
            table_data[1][2] = img_obj



        # --- Create table ---
        table = Table(
            table_data,
            colWidths=[table_width * 0.25, table_width * 0.45, table_width * 0.3],
            rowHeights=row_heights,
        )
        
        # --- Table style ---
        style = [
            ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
            # Center all label cells (col 0)
            ("ALIGN", (0, 0), (0, len(table_data)-1), "CENTER"),
            ("VALIGN", (0, 0), (0, len(table_data)-1), "MIDDLE"),
            # Center PHOTO header
            ("ALIGN", (2, 0), (2, 0), "CENTER"),
            ("VALIGN", (2, 0), (2, 0), "MIDDLE"),
            # Center image (col 2, all rows except header)
            ("ALIGN", (2, 1), (2, len(table_data)-2), "CENTER"),
            ("VALIGN", (2, 1), (2, len(table_data)-2), "MIDDLE"),
            # Merge image column
            ("SPAN", (2, 1), (2, len(table_data)-2)),
            # Merge description across 2nd+3rd columns
            ("SPAN", (1, 8), (2, 8)),
            # Value cells (col 1) left-aligned, top-aligned
            ("ALIGN", (1, 0), (1, len(table_data)-1), "LEFT"),
            ("VALIGN", (1, 0), (1, len(table_data)-2), "MIDDLE"),
            ("VALIGN", (1, 8), (1, 8), "TOP"),  # description value top-aligned
        ]

        table.setStyle(TableStyle(style))
        return table

    # --- Build PDF ---
    for i, artefact in enumerate(artefacts):
        table = build_table(artefact)
        elements.append(table)
        elements.append(Spacer(1, 40))
        if i % 2 == 1:
            elements.append(PageBreak())

    doc.build(elements)
    print(f"✅ Exported to PDF: {filename}")